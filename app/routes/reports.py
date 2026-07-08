import os, io, tempfile
from datetime import date, timedelta, datetime
from collections import defaultdict
from flask import (Blueprint, render_template, request, jsonify,
                   send_file, current_app)
from flask_login import login_required, current_user
from sqlalchemy import func, text

from ..extensions import db
from ..services.mailer import send_mail
from ..models.sales    import SalesMaster, SalesItem, Return
from ..models.product  import Product
from ..models.customer import Customer

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")

# ── Shared profit SQL fragment ─────────────────────────────────
_PROFIT_SQL = """
    (si.mrp - IFNULL(p.cost_price,0)) * si.quantity
    - CASE si.discount_type
        WHEN 'Flat' THEN IFNULL(si.discount_value,0)
        ELSE si.mrp * si.quantity * IFNULL(si.discount_value,0) / 100.0
      END
"""

_DISC_SQL = """
    CASE si.discount_type
        WHEN 'Flat' THEN IFNULL(si.discount_value,0)
        ELSE si.mrp * si.quantity * IFNULL(si.discount_value,0) / 100.0
    END
"""


def _admin_only():
    if current_user.role != "Admin":
        return jsonify(error="Admin access required"), 403


def _date_range():
    s = request.args.get("from", (date.today() - timedelta(days=30)).isoformat())
    e = request.args.get("to",   date.today().isoformat())
    return s, e


def _wants_email():
    """?email=1 / true / yes on any export route routes it to inbox instead of download."""
    return request.args.get("email", "").strip().lower() in ("1", "true", "yes")


def _email_report(filename, mimetype, file_bytes, subject, body):
    """
    Emails a generated report file to the fixed admin inbox
    (ADMIN_REPORT_EMAIL) via Mailjet instead of streaming it as a download.
    Returns a Flask jsonify response.
    """
    recipient = current_app.config.get("ADMIN_REPORT_EMAIL")
    if not recipient:
        return jsonify(success=False,
                       message="ADMIN_REPORT_EMAIL is not configured."), 400
    try:
        send_mail(
            subject=subject,
            recipients=[recipient],
            body=body,
            attachments=[(filename, mimetype, file_bytes)],
        )
        return jsonify(success=True, message=f"Report emailed to {recipient}")
    except Exception as ex:
        return jsonify(success=False, message=f"Email failed: {ex}"), 500


# ══════════════════════════════════════════════════════════════
# PAGE
# ══════════════════════════════════════════════════════════════

@reports_bp.route("/")
@login_required
def index():
    if current_user.role != "Admin":
        from flask import flash, redirect, url_for
        flash("Reports are Admin only.", "warning")
        return redirect(url_for("dashboard.home"))
    return render_template("reports/index.html")


# ══════════════════════════════════════════════════════════════
# KPIs
# ══════════════════════════════════════════════════════════════

@reports_bp.route("/api/kpis")
@login_required
def api_kpis():
    err = _admin_only()
    if err: return err

    s, e = _date_range()

    total_sales = db.session.execute(
        text("SELECT IFNULL(SUM(grand_total),0) FROM sales_master WHERE DATE(date) BETWEEN :s AND :e"),
        {"s": s, "e": e}
    ).scalar() or 0.0

    total_profit = db.session.execute(text(f"""
        SELECT IFNULL(SUM({_PROFIT_SQL}),0)
        FROM sales_items si
        LEFT JOIN products p ON si.product_id=p.product_id
        JOIN sales_master sm ON si.sale_id=sm.sale_id
        WHERE DATE(sm.date) BETWEEN :s AND :e
    """), {"s": s, "e": e}).scalar() or 0.0

    avg_row = db.session.execute(text("""
        SELECT IFNULL(SUM(grand_total),0) AS tot,
               COUNT(DISTINCT DATE(date)) AS days
        FROM sales_master
        WHERE DATE(date) BETWEEN :s AND :e
    """), {"s": s, "e": e}).fetchone()
    days      = max(1, avg_row[1] or 1)
    avg_daily = float(avg_row[0] or 0) / days

    return_val = db.session.execute(
        text("SELECT IFNULL(SUM(refund_amount),0) FROM returns WHERE DATE(date) BETWEEN :s AND :e"),
        {"s": s, "e": e}
    ).scalar() or 0.0

    bill_count = db.session.execute(
        text("SELECT COUNT(*) FROM sales_master WHERE DATE(date) BETWEEN :s AND :e"),
        {"s": s, "e": e}
    ).scalar() or 0

    customers = Customer.query.count()
    margin    = (float(total_profit) / float(total_sales) * 100 if total_sales else 0.0)

    return jsonify(
        total_sales=round(float(total_sales), 2),
        net_profit=round(float(total_profit), 2),
        margin=round(margin, 2),
        avg_daily=round(avg_daily, 2),
        total_returns=round(float(return_val), 2),
        customers=customers,
        bill_count=int(bill_count),
    )


# ══════════════════════════════════════════════════════════════
# SALES HISTORY
# ══════════════════════════════════════════════════════════════

@reports_bp.route("/api/sales-history")
@login_required
def api_sales_history():
    err = _admin_only()
    if err: return err
    s, e     = _date_range()
    page     = request.args.get("page",     1,    type=int)
    per      = min(request.args.get("per_page", 100, type=int), 500)
    sold_by  = request.args.get("sold_by",  "").strip()
    category = request.args.get("category", "").strip()
    pay_mode = request.args.get("payment_mode", "").strip()

    filters  = "WHERE DATE(sm.date) BETWEEN :s AND :e"
    params   = {"s": s, "e": e}
    if sold_by:
        filters += " AND sm.sold_by = :sold_by"
        params["sold_by"] = sold_by
    if category:
        filters += " AND si.category = :category"
        params["category"] = category
    if pay_mode:
        filters += " AND sm.payment_mode LIKE :pay_mode"
        params["pay_mode"] = f"%{pay_mode}%"

    base_sql = f"""
        FROM sales_master sm
        JOIN sales_items si ON si.sale_id=sm.sale_id
        LEFT JOIN products p ON si.product_id=p.product_id
        {filters}
    """

    summary_row = db.session.execute(text(f"""
        SELECT COUNT(DISTINCT sm.sale_id)           AS bill_count,
               IFNULL(SUM(si.effective_total),0)    AS total_revenue,
               IFNULL(SUM(IFNULL(p.gst,0)/100.0 * (
                   si.mrp*si.quantity
                   - CASE si.discount_type WHEN 'Flat'
                       THEN IFNULL(si.discount_value,0)
                       ELSE si.mrp*si.quantity*IFNULL(si.discount_value,0)/100.0
                     END
               )),0)                                AS total_gst,
               IFNULL(SUM({_DISC_SQL}),0)           AS total_discount,
               IFNULL(SUM({_PROFIT_SQL}),0)         AS total_profit,
               COUNT(*)                             AS row_count
        {base_sql}
    """), params).fetchone()

    offset = (page - 1) * per
    params["limit"]  = per
    params["offset"] = offset

    rows = db.session.execute(text(f"""
        SELECT sm.sale_id, sm.date, sm.payment_mode,
               si.product_name, si.category,
               si.quantity, si.mrp AS unit_price,
               IFNULL(p.cost_price,0) AS cost_price,
               si.discount_type, si.discount_value,
               si.effective_total, IFNULL(p.gst,0) AS gst_pct,
               sm.sold_by, sm.customer_name, sm.customer_phone
        {base_sql}
        ORDER BY sm.date DESC
        LIMIT :limit OFFSET :offset
    """), params).fetchall()

    result = []
    for r in rows:
        qty      = int(r.quantity or 1)
        unit_p   = float(r.unit_price  or 0)
        cost_p   = float(r.cost_price  or 0)
        dt_      = r.discount_type or "Flat"
        dv       = float(r.discount_value or 0)
        gst_pct  = float(r.gst_pct or 0)
        disc_amt = dv if dt_ == "Flat" else unit_p * qty * dv / 100.0
        net_prof = round((unit_p - cost_p) * qty - disc_amt, 2)
        taxable  = max(unit_p * qty - disc_amt, 0.0)
        gst_amt  = round(taxable * gst_pct / 100.0, 2)
        margin   = (net_prof / (unit_p * qty) * 100) if unit_p * qty else 0.0
        result.append({
            "sale_id":        r.sale_id,
            "date":           str(r.date),
            "product_name":   r.product_name,
            "category":       r.category or "",
            "qty":            qty,
            "unit_price":     unit_p,
            "cost_price":     cost_p,
            "discount":       f"Flat ₹{dv:.2f}" if dt_ == "Flat" else f"{dv:.0f}%",
            "net_profit":     net_prof,
            "margin_pct":     round(margin, 2),
            "gst":            gst_amt,
            "total":          float(r.effective_total or 0),
            "sold_by":        r.sold_by,
            "customer":       r.customer_name,
            "customer_phone": r.customer_phone or "",
            "payment_mode":   r.payment_mode or "",
        })

    return jsonify(
        rows=result,
        page=page,
        per_page=per,
        summary={
            "bill_count":     int(summary_row.bill_count   or 0),
            "total_revenue":  round(float(summary_row.total_revenue  or 0), 2),
            "total_gst":      round(float(summary_row.total_gst      or 0), 2),
            "total_discount": round(float(summary_row.total_discount or 0), 2),
            "total_profit":   round(float(summary_row.total_profit   or 0), 2),
            "row_count":      int(summary_row.row_count    or 0),
        }
    )


# ══════════════════════════════════════════════════════════════
# PROFIT REPORT
# ══════════════════════════════════════════════════════════════

@reports_bp.route("/api/profit")
@login_required
def api_profit():
    err = _admin_only()
    if err: return err
    s, e     = _date_range()
    group_by = request.args.get("group_by", "product")

    try:
        if group_by == "category":
            group_col   = "si.category"
            select_name = "IFNULL(si.category,'Other') AS label"
        elif group_by == "supplier":
            group_col   = "p.supplier_id"
            select_name = "IFNULL(p.supplier_id,'Unknown') AS label"
        else:
            group_col   = "si.product_id, si.product_name"
            select_name = "si.product_name AS label"

        rows = db.session.execute(text(f"""
            SELECT {select_name},
                   SUM(si.quantity) AS qty_sold,
                   IFNULL(SUM(si.effective_total),0) AS revenue,
                   IFNULL(SUM({_DISC_SQL}),0) AS total_disc,
                   IFNULL(SUM(IFNULL(p.cost_price,0)*si.quantity),0) AS cogs,
                   IFNULL(SUM({_PROFIT_SQL}),0) AS net_profit
            FROM sales_items si
            LEFT JOIN products p ON si.product_id=p.product_id
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE DATE(sm.date) BETWEEN :s AND :e
            GROUP BY {group_col}
            ORDER BY revenue DESC
        """), {"s": s, "e": e}).fetchall()

        real_total_rev = sum(float(r.revenue or 0) for r in rows)
        total_rev      = real_total_rev or 1
        cumulative     = 0.0
        result         = []
        total_profit   = 0.0

        for r in rows:
            rev    = float(r.revenue    or 0)
            profit = float(r.net_profit or 0)
            disc   = float(r.total_disc or 0)
            cogs   = float(r.cogs       or 0)
            margin = (profit / rev * 100) if rev else 0.0
            cumulative   += rev / total_rev * 100
            total_profit += profit
            abc = "A" if cumulative <= 70 else "B" if cumulative <= 90 else "C"
            result.append({
                "label":      r.label,
                "qty_sold":   int(r.qty_sold or 0),
                "revenue":    round(rev, 2),
                "total_disc": round(disc, 2),
                "cogs":       round(cogs, 2),
                "net_profit": round(profit, 2),
                "margin_pct": round(margin, 2),
                "abc":        abc,
            })

        days = max(1, (date.fromisoformat(e) - date.fromisoformat(s)).days + 1)
        return jsonify(
            success=True,
            group_by=group_by,
            rows=result,
            summary={
                "total_revenue":    round(real_total_rev, 2),
                "net_profit":       round(total_profit, 2),
                "margin":           round(total_profit / real_total_rev * 100, 2) if real_total_rev else 0.0,
                "avg_daily_profit": round(total_profit / days, 2),
            }
        )
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


# ══════════════════════════════════════════════════════════════
# GST REPORT
# ══════════════════════════════════════════════════════════════

@reports_bp.route("/api/gst")
@login_required
def api_gst():
    err = _admin_only()
    if err: return err
    s, e = _date_range()

    try:
        rows = db.session.execute(text(f"""
            SELECT si.product_name, si.category,
                   IFNULL(p.gst,0) AS gst_rate,
                   SUM(si.quantity) AS qty,
                   COUNT(DISTINCT sm.sale_id) AS txns,
                   IFNULL(SUM(
                       si.mrp*si.quantity - ({_DISC_SQL})
                   ),0) AS taxable_amt
            FROM sales_items si
            LEFT JOIN products p ON si.product_id=p.product_id
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE DATE(sm.date) BETWEEN :s AND :e
            GROUP BY si.product_id, si.product_name, si.category, p.gst
            ORDER BY taxable_amt DESC
        """), {"s": s, "e": e}).fetchall()

        cat_map   = {}
        rate_map  = {}
        prod_data = []

        for r in rows:
            gst_pct = float(r.gst_rate or 0)
            taxable = max(float(r.taxable_amt or 0), 0.0)
            half    = gst_pct / 2.0
            cgst    = round(taxable * half / 100.0, 2)
            sgst    = round(taxable * half / 100.0, 2)

            prod_data.append({
                "product":     r.product_name,
                "category":    r.category or "Other",
                "gst_rate":    gst_pct,
                "qty":         int(r.qty or 0),
                "taxable_amt": round(taxable, 2),
                "cgst":        cgst,
                "sgst":        sgst,
                "total_gst":   round(cgst + sgst, 2),
            })

            cat = r.category or "Other"
            if cat not in cat_map:
                cat_map[cat] = {"rate": gst_pct, "txns": 0, "taxable": 0.0, "cgst": 0.0, "sgst": 0.0}
            cat_map[cat]["txns"]    += int(r.txns or 0)
            cat_map[cat]["taxable"] += taxable
            cat_map[cat]["cgst"]    += cgst
            cat_map[cat]["sgst"]    += sgst

            rate_key = int(gst_pct)
            if rate_key not in rate_map:
                rate_map[rate_key] = {"taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "txns": 0}
            rate_map[rate_key]["taxable"] += taxable
            rate_map[rate_key]["cgst"]    += cgst
            rate_map[rate_key]["sgst"]    += sgst
            rate_map[rate_key]["txns"]    += int(r.txns or 0)

        cat_rows = []
        grand_t = grand_c = grand_s = 0.0
        for cat, v in sorted(cat_map.items(), key=lambda x: -x[1]["taxable"]):
            grand_t += v["taxable"]
            grand_c += v["cgst"]
            grand_s += v["sgst"]
            cat_rows.append({
                "category":     cat,
                "gst_rate":     v["rate"],
                "transactions": v["txns"],
                "taxable_amt":  round(v["taxable"], 2),
                "cgst":         round(v["cgst"], 2),
                "sgst":         round(v["sgst"], 2),
                "total_gst":    round(v["cgst"] + v["sgst"], 2),
            })

        by_rate_rows = []
        for rate in sorted(rate_map.keys()):
            v = rate_map[rate]
            by_rate_rows.append({
                "gst_rate":     rate,
                "transactions": v["txns"],
                "taxable_amt":  round(v["taxable"], 2),
                "cgst":         round(v["cgst"], 2),
                "sgst":         round(v["sgst"], 2),
                "total_gst":    round(v["cgst"] + v["sgst"], 2),
            })

        return jsonify(
            success=True,
            products=prod_data,
            categories=cat_rows,
            by_rate=by_rate_rows,
            summary={
                "period_from": s,
                "period_to":   e,
                "taxable":     round(grand_t, 2),
                "cgst":        round(grand_c, 2),
                "sgst":        round(grand_s, 2),
                "total_gst":   round(grand_c + grand_s, 2),
            }
        )
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/gstr1")
@login_required
def api_gstr1():
    err = _admin_only()
    if err: return err
    month = request.args.get("month", date.today().strftime("%m"))
    year  = request.args.get("year",  date.today().strftime("%Y"))
    try:
        rows = db.session.execute(text(f"""
            SELECT IFNULL(p.gst,0) AS gst_rate,
                   COUNT(DISTINCT sm.sale_id) AS txns,
                   SUM(si.quantity)           AS total_qty,
                   IFNULL(SUM(
                       si.mrp*si.quantity - ({_DISC_SQL})
                   ),0) AS taxable_amt
            FROM sales_items si
            LEFT JOIN products p ON si.product_id=p.product_id
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE MONTH(sm.date)=:month AND YEAR(sm.date)=:year
            GROUP BY p.gst
            ORDER BY p.gst
        """), {"month": int(month), "year": int(year)}).fetchall()

        slabs = []
        grand_taxable = grand_cgst = grand_sgst = 0.0
        for r in rows:
            rate    = float(r.gst_rate or 0)
            taxable = max(float(r.taxable_amt or 0), 0.0)
            cgst    = round(taxable * (rate/2) / 100, 2)
            sgst    = round(taxable * (rate/2) / 100, 2)
            grand_taxable += taxable
            grand_cgst    += cgst
            grand_sgst    += sgst
            slabs.append({
                "gst_rate":     rate,
                "transactions": int(r.txns or 0),
                "total_qty":    int(r.total_qty or 0),
                "taxable_amt":  round(taxable, 2),
                "cgst":         cgst,
                "sgst":         sgst,
                "total_tax":    round(cgst + sgst, 2),
            })

        return jsonify(
            success=True,
            period=f"{year}-{month}",
            slabs=slabs,
            totals={
                "taxable":   round(grand_taxable, 2),
                "cgst":      round(grand_cgst, 2),
                "sgst":      round(grand_sgst, 2),
                "total_tax": round(grand_cgst + grand_sgst, 2),
            }
        )
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


# ══════════════════════════════════════════════════════════════
# CHARTS
# ══════════════════════════════════════════════════════════════

@reports_bp.route("/api/chart/monthly")
@login_required
def api_chart_monthly():
    err = _admin_only()
    if err: return err
    try:
        rows = db.session.execute(text(f"""
            SELECT DATE_FORMAT(sm.date,'%Y-%m') AS ym,
                   IFNULL(SUM(sm.grand_total),0) AS sales,
                   IFNULL(SUM({_PROFIT_SQL}),0)  AS profit,
                   COUNT(DISTINCT sm.sale_id)     AS bill_count
            FROM sales_master sm
            JOIN sales_items si ON si.sale_id=sm.sale_id
            LEFT JOIN products p ON si.product_id=p.product_id
            GROUP BY ym ORDER BY ym DESC LIMIT 12
        """)).fetchall()
        rows = list(reversed(rows))
        return jsonify(success=True, labels=[r[0] for r in rows],
                       sales=[float(r[1] or 0) for r in rows],
                       profit=[float(r[2] or 0) for r in rows],
                       bill_count=[int(r[3] or 0) for r in rows])
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/chart/daily")
@login_required
def api_chart_daily():
    err = _admin_only()
    if err: return err
    try:
        days_back = int(request.args.get("days", 30))
        start     = (date.today() - timedelta(days=days_back - 1)).isoformat()
        rows = db.session.execute(text(f"""
            SELECT DATE(sm.date) AS day,
                   IFNULL(SUM(sm.grand_total),0) AS sales,
                   IFNULL(SUM({_PROFIT_SQL}),0)  AS profit
            FROM sales_master sm
            JOIN sales_items si ON si.sale_id=sm.sale_id
            LEFT JOIN products p ON si.product_id=p.product_id
            WHERE DATE(sm.date) >= :start
            GROUP BY day ORDER BY day
        """), {"start": start}).fetchall()
        sales_vals = [float(r[1] or 0) for r in rows]
        cumulative = []
        running = 0.0
        for v in sales_vals:
            running += v
            cumulative.append(round(running, 2))
        return jsonify(success=True, labels=[str(r[0]) for r in rows],
                       sales=sales_vals, profit=[float(r[2] or 0) for r in rows],
                       cumulative=cumulative)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/chart/top-products")
@login_required
def api_chart_top():
    err = _admin_only()
    if err: return err
    sort_by = request.args.get("sort_by", "profit")
    limit   = min(int(request.args.get("limit", 10)), 20)
    s, e    = _date_range()
    order_col = {"revenue": "revenue", "qty": "qty_sold"}.get(sort_by, "profit")
    try:
        rows = db.session.execute(text(f"""
            SELECT si.product_name,
                   IFNULL(SUM(si.effective_total),0) AS revenue,
                   IFNULL(SUM({_PROFIT_SQL}),0)      AS profit,
                   SUM(si.quantity)                  AS qty_sold
            FROM sales_items si
            LEFT JOIN products p ON si.product_id=p.product_id
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE DATE(sm.date) BETWEEN :s AND :e
            GROUP BY si.product_id, si.product_name
            ORDER BY {order_col} DESC LIMIT :lim
        """), {"s": s, "e": e, "lim": limit}).fetchall()
        return jsonify(success=True, sort_by=sort_by,
                       labels=[r[0] for r in rows],
                       sales=[float(r[1] or 0) for r in rows],
                       profit=[float(r[2] or 0) for r in rows],
                       qty=[int(r[3] or 0) for r in rows])
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/chart/category-mix")
@login_required
def api_chart_category():
    err = _admin_only()
    if err: return err
    s, e = _date_range()
    try:
        rows = db.session.execute(text(f"""
            SELECT IFNULL(si.category,'Other') AS category,
                   IFNULL(SUM(si.effective_total),0)         AS revenue,
                   IFNULL(SUM({_PROFIT_SQL}),0)              AS profit,
                   IFNULL(SUM(IFNULL(p.cost_price,0)*si.quantity),0) AS cogs
            FROM sales_items si
            LEFT JOIN products p ON si.product_id=p.product_id
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE DATE(sm.date) BETWEEN :s AND :e
            GROUP BY si.category ORDER BY revenue DESC
        """), {"s": s, "e": e}).fetchall()
        labels  = [r[0] or "Other" for r in rows]
        revenue = [float(r[1] or 0) for r in rows]
        profit  = [float(r[2] or 0) for r in rows]
        cogs    = [float(r[3] or 0) for r in rows]
        margin  = [round(p / r * 100, 2) if r else 0.0 for p, r in zip(profit, revenue)]
        return jsonify(success=True, labels=labels, revenue=revenue,
                       profit=profit, cogs=cogs, margin=margin)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/chart/weekly-pattern")
@login_required
def api_chart_weekly_pattern():
    err = _admin_only()
    if err: return err
    s, e = _date_range()
    try:
        rows = db.session.execute(text("""
            SELECT DAYOFWEEK(sm.date) AS dow,
                   IFNULL(AVG(daily.day_total),0) AS avg_sales,
                   IFNULL(SUM(daily.day_total),0) AS total_sales,
                   COUNT(DISTINCT DATE(sm.date))  AS day_count
            FROM sales_master sm
            JOIN (
                SELECT DATE(date) AS d, SUM(grand_total) AS day_total
                FROM sales_master
                WHERE DATE(date) BETWEEN :s AND :e
                GROUP BY DATE(date)
            ) daily ON DATE(sm.date) = daily.d
            WHERE DATE(sm.date) BETWEEN :s AND :e
            GROUP BY dow ORDER BY dow
        """), {"s": s, "e": e}).fetchall()
        day_names = {1:"Sun",2:"Mon",3:"Tue",4:"Wed",5:"Thu",6:"Fri",7:"Sat"}
        data = {r[0]: {"avg": float(r[1] or 0), "total": float(r[2] or 0)} for r in rows}
        labels = [day_names[d] for d in range(1,8)]
        avg_vals = [round(data.get(d,{}).get("avg",0.0),2) for d in range(1,8)]
        totals   = [round(data.get(d,{}).get("total",0.0),2) for d in range(1,8)]
        return jsonify(success=True, labels=labels, avg_sales=avg_vals, total_sales=totals)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/chart/hourly")
@login_required
def api_chart_hourly():
    err = _admin_only()
    if err: return err
    s, e = _date_range()
    try:
        rows = db.session.execute(text("""
            SELECT HOUR(sm.date) AS hr,
                   IFNULL(SUM(sm.grand_total),0) AS total_sales,
                   COUNT(DISTINCT sm.sale_id)     AS bill_count
            FROM sales_master sm
            WHERE DATE(sm.date) BETWEEN :s AND :e
            GROUP BY hr ORDER BY hr
        """), {"s": s, "e": e}).fetchall()
        data        = {int(r[0]): {"sales": float(r[1] or 0), "bills": int(r[2] or 0)} for r in rows}
        labels      = [f"{h:02d}:00" for h in range(24)]
        sales_vals  = [round(data.get(h,{}).get("sales",0.0),2) for h in range(24)]
        bill_counts = [data.get(h,{}).get("bills",0) for h in range(24)]
        return jsonify(success=True, labels=labels, sales=sales_vals, bill_count=bill_counts)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/chart/payment-mix")
@login_required
def api_chart_payment_mix():
    err = _admin_only()
    if err: return err
    s, e = _date_range()
    try:
        rows = db.session.execute(text("""
            SELECT payment_mode, COUNT(*) AS bill_count,
                   IFNULL(SUM(grand_total),0) AS total_value
            FROM sales_master
            WHERE DATE(date) BETWEEN :s AND :e
            GROUP BY payment_mode ORDER BY total_value DESC
        """), {"s": s, "e": e}).fetchall()
        return jsonify(success=True,
                       labels=[r[0] or "Unknown" for r in rows],
                       bill_count=[int(r[1] or 0) for r in rows],
                       total_value=[float(r[2] or 0) for r in rows])
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/chart/margin-heatmap")
@login_required
def api_chart_margin_heatmap():
    err = _admin_only()
    if err: return err
    months = min(int(request.args.get("months", 6)), 12)
    try:
        rows = db.session.execute(text(f"""
            SELECT IFNULL(si.category,'Other')      AS category,
                   DATE_FORMAT(sm.date,'%Y-%m')      AS ym,
                   IFNULL(SUM(si.effective_total),0) AS revenue,
                   IFNULL(SUM({_PROFIT_SQL}),0)      AS profit
            FROM sales_items si
            LEFT JOIN products p ON si.product_id=p.product_id
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE sm.date >= DATE_SUB(NOW(), INTERVAL :months MONTH)
            GROUP BY si.category, ym ORDER BY si.category, ym
        """), {"months": months}).fetchall()
        matrix     = defaultdict(dict)
        all_months = set()
        for r in rows:
            rev    = float(r.revenue or 0)
            profit = float(r.profit  or 0)
            margin = round(profit / rev * 100, 1) if rev else 0.0
            matrix[r.category][r.ym] = margin
            all_months.add(r.ym)
        month_labels = sorted(all_months)
        categories   = sorted(matrix.keys())
        grid = [[matrix[cat].get(m, None) for m in month_labels] for cat in categories]
        return jsonify(success=True, categories=categories, months=month_labels, grid=grid)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


# ══════════════════════════════════════════════════════════════
# ADVANCED REPORTS
# ══════════════════════════════════════════════════════════════

@reports_bp.route("/api/dead-stock")
@login_required
def api_dead_stock():
    err = _admin_only()
    if err: return err
    try:
        days_param = int(request.args.get("days", 30))
        category   = request.args.get("category", "").strip()
        cutoff     = (date.today() - timedelta(days=days_param)).isoformat()
        cat_filter = ""; params = {"cutoff": cutoff}
        if category: cat_filter = "AND p.category = :cat"; params["cat"] = category

        rows = db.session.execute(text(f"""
            SELECT p.product_id, p.name, p.category,
                   p.quantity, IFNULL(p.reorder_level,0) AS rl,
                   IFNULL(p.cost_price,0) AS cp,
                   p.quantity*IFNULL(p.cost_price,0) AS stock_val,
                   (SELECT MAX(DATE(sm2.date)) FROM sales_items si2
                    JOIN sales_master sm2 ON si2.sale_id=sm2.sale_id
                    WHERE si2.product_id=p.product_id) AS last_sold_date
            FROM products p
            WHERE p.quantity > 0 {cat_filter}
              AND p.product_id NOT IN (
                  SELECT DISTINCT si.product_id FROM sales_items si
                  JOIN sales_master sm ON si.sale_id=sm.sale_id
                  WHERE DATE(sm.date) >= :cutoff)
            ORDER BY stock_val DESC
        """), params).fetchall()

        data = [{"product_id":r[0],"name":r[1],"category":r[2] or "","qty":r[3],"reorder":r[4],
                 "cost":float(r[5] or 0),"stock_val":float(r[6] or 0),
                 "last_sold_date":str(r[7]) if r[7] else "Never",
                 "days_inactive":(date.today()-date.fromisoformat(str(r[7]))).days if r[7] else None}
                for r in rows]
        return jsonify(success=True, days_threshold=days_param, rows=data,
                       total_locked=round(sum(r["stock_val"] for r in data),2))
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/return-rate")
@login_required
def api_return_rate():
    err = _admin_only()
    if err: return err
    s, e = _date_range()
    limit = min(int(request.args.get("limit", 15)), 50)
    category = request.args.get("category", "").strip()
    cat_filter = ""; params = {"s":s,"e":e,"lim":limit}
    if category: cat_filter = "AND si.category = :cat"; params["cat"] = category
    try:
        rows = db.session.execute(text(f"""
            SELECT si.product_name, SUM(si.quantity) AS sold,
                   IFNULL(r.ret_qty,0) AS returned, si.category
            FROM sales_items si
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            LEFT JOIN (SELECT product_id, SUM(quantity) AS ret_qty FROM returns GROUP BY product_id) r
                ON r.product_id=si.product_id
            WHERE DATE(sm.date) BETWEEN :s AND :e {cat_filter}
            GROUP BY si.product_id, si.product_name, si.category HAVING sold > 0
            ORDER BY (CAST(IFNULL(r.ret_qty,0) AS DECIMAL)/SUM(si.quantity)) DESC LIMIT :lim
        """), params).fetchall()
        return jsonify(success=True, rows=[{"product":r[0],"sold":int(r[1] or 0),
            "returned":int(r[2] or 0),"category":r[3] or "",
            "rate":round(int(r[2] or 0)/int(r[1])*100,2) if r[1] else 0.0} for r in rows])
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/mom-comparison")
@login_required
def api_mom_comparison():
    err = _admin_only()
    if err: return err
    months = min(int(request.args.get("months", 6)), 12)
    try:
        rows = db.session.execute(text(f"""
            SELECT DATE_FORMAT(sm.date,'%Y-%m') AS ym,
                   IFNULL(SUM(sm.grand_total),0)  AS total_sales,
                   IFNULL(SUM({_PROFIT_SQL}),0)   AS total_profit,
                   COUNT(DISTINCT sm.sale_id)      AS bill_count,
                   IFNULL(SUM({_DISC_SQL}),0)      AS total_discount,
                   IFNULL(SUM(r_sub.refund),0)     AS total_returns
            FROM sales_master sm
            JOIN sales_items si ON si.sale_id=sm.sale_id
            LEFT JOIN products p ON si.product_id=p.product_id
            LEFT JOIN (SELECT DATE_FORMAT(date,'%Y-%m') AS rym, SUM(refund_amount) AS refund
                       FROM returns GROUP BY rym) r_sub ON r_sub.rym=DATE_FORMAT(sm.date,'%Y-%m')
            WHERE sm.date >= DATE_SUB(NOW(), INTERVAL :months MONTH)
            GROUP BY ym ORDER BY ym ASC
        """), {"months": months}).fetchall()

        result=[]; prev=None
        for r in rows:
            sales=round(float(r.total_sales or 0),2); profit=round(float(r.total_profit or 0),2)
            bills=int(r.bill_count or 0)
            discount=round(float(r.total_discount or 0),2); returns=round(float(r.total_returns or 0),2)
            avg_bill=round(sales/bills,2) if bills else 0.0; margin=round(profit/sales*100,2) if sales else 0.0
            pct=lambda c,p:round((c-p)/p*100,1) if p else None
            row={"month":r.ym,"total_sales":sales,"total_profit":profit,"bill_count":bills,
                 "avg_bill":avg_bill,"margin_pct":margin,"total_discount":discount,"total_returns":returns,
                 "vs_prev":None if prev is None else
                     {"sales_chg":pct(sales,prev["total_sales"]),"profit_chg":pct(profit,prev["total_profit"]),
                      "bills_chg":pct(bills,prev["bill_count"])}}
            result.append(row); prev=row
        return jsonify(success=True, months=result)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/staff-performance")
@login_required
def api_staff_performance():
    err = _admin_only()
    if err: return err
    s, e = _date_range()
    try:
        rows = db.session.execute(text(f"""
            SELECT sm.sold_by, COUNT(DISTINCT sm.sale_id) AS bill_count,
                   IFNULL(SUM(sm.grand_total),0) AS total_revenue,
                   IFNULL(SUM({_PROFIT_SQL}),0)  AS total_profit,
                   IFNULL(SUM({_DISC_SQL}),0)     AS total_discount,
                   COUNT(DISTINCT ret.return_id)  AS returns_processed
            FROM sales_master sm
            JOIN sales_items si ON si.sale_id=sm.sale_id
            LEFT JOIN products p ON si.product_id=p.product_id
            LEFT JOIN returns ret ON ret.sale_id=sm.sale_id
            WHERE DATE(sm.date) BETWEEN :s AND :e
            GROUP BY sm.sold_by ORDER BY total_revenue DESC
        """), {"s":s,"e":e}).fetchall()
        result=[]
        for r in rows:
            rev=float(r.total_revenue or 0); bills=int(r.bill_count or 0)
            result.append({"employee":r.sold_by,"bill_count":bills,"total_revenue":round(rev,2),
                "total_profit":round(float(r.total_profit or 0),2),
                "total_discount":round(float(r.total_discount or 0),2),
                "avg_bill":round(rev/bills,2) if bills else 0.0,
                "returns_processed":int(r.returns_processed or 0)})
        return jsonify(success=True, rows=result)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/customer-sales")
@login_required
def api_customer_sales():
    err = _admin_only()
    if err: return err
    s, e = _date_range()
    try:
        rows = db.session.execute(text("""
            SELECT sm.customer_name, IFNULL(sm.customer_phone,'') AS phone,
                   COUNT(DISTINCT sm.sale_id) AS order_count,
                   IFNULL(SUM(sm.grand_total),0) AS total_spend,
                   MIN(DATE(sm.date)) AS first_purchase,
                   MAX(DATE(sm.date)) AS last_purchase
            FROM sales_master sm
            WHERE DATE(sm.date) BETWEEN :s AND :e
            GROUP BY sm.customer_name, sm.customer_phone
            ORDER BY total_spend DESC
        """), {"s":s,"e":e}).fetchall()
        result=[]
        for r in rows:
            spend=float(r.total_spend or 0); orders=int(r.order_count or 0)
            result.append({"customer":r.customer_name,"phone":r.phone or "","order_count":orders,
                "total_spend":round(spend,2),"avg_order":round(spend/orders,2) if orders else 0.0,
                "first_purchase":str(r.first_purchase),"last_purchase":str(r.last_purchase)})
        return jsonify(success=True, rows=result)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/discount-analysis")
@login_required
def api_discount_analysis():
    err = _admin_only()
    if err: return err
    s, e = _date_range()
    try:
        by_product = db.session.execute(text(f"""
            SELECT si.product_name, SUM(si.quantity) AS qty_sold,
                   IFNULL(SUM(si.mrp*si.quantity),0) AS gross_revenue,
                   IFNULL(SUM({_DISC_SQL}),0)        AS total_discount,
                   IFNULL(SUM(si.effective_total),0) AS net_revenue
            FROM sales_items si JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE DATE(sm.date) BETWEEN :s AND :e
              AND si.discount_value IS NOT NULL AND si.discount_value > 0
            GROUP BY si.product_id, si.product_name ORDER BY total_discount DESC LIMIT 20
        """), {"s":s,"e":e}).fetchall()

        by_employee = db.session.execute(text(f"""
            SELECT sm.sold_by, COUNT(DISTINCT sm.sale_id) AS discounted_bills,
                   IFNULL(SUM({_DISC_SQL}),0)        AS total_discount,
                   IFNULL(SUM(si.effective_total),0) AS net_revenue
            FROM sales_items si JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE DATE(sm.date) BETWEEN :s AND :e
              AND si.discount_value IS NOT NULL AND si.discount_value > 0
            GROUP BY sm.sold_by ORDER BY total_discount DESC
        """), {"s":s,"e":e}).fetchall()

        return jsonify(success=True,
            by_product=[{"product":r.product_name,"qty_sold":int(r.qty_sold or 0),
                "gross_revenue":round(float(r.gross_revenue or 0),2),
                "total_discount":round(float(r.total_discount or 0),2),
                "net_revenue":round(float(r.net_revenue or 0),2),
                "discount_pct":round(float(r.total_discount or 0)/float(r.gross_revenue)*100,2)
                    if r.gross_revenue else 0.0} for r in by_product],
            by_employee=[{"employee":r.sold_by,"discounted_bills":int(r.discounted_bills or 0),
                "total_discount":round(float(r.total_discount or 0),2),
                "net_revenue":round(float(r.net_revenue or 0),2),
                "discount_pct":round(float(r.total_discount or 0)/(float(r.net_revenue or 0)+float(r.total_discount or 0))*100,2)
                    if (r.net_revenue or 0)+(r.total_discount or 0) else 0.0} for r in by_employee])
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/sales-velocity")
@login_required
def api_sales_velocity():
    err = _admin_only()
    if err: return err
    days_param = int(request.args.get("days", 30))
    start      = (date.today() - timedelta(days=days_param)).isoformat()
    try:
        rows = db.session.execute(text("""
            SELECT p.product_id, p.name, p.category,
                   p.quantity AS current_stock, IFNULL(SUM(si.quantity),0) AS sold_qty
            FROM products p
            LEFT JOIN sales_items si ON si.product_id=p.product_id
            LEFT JOIN sales_master sm ON si.sale_id=sm.sale_id AND DATE(sm.date) >= :start
            GROUP BY p.product_id, p.name, p.category, p.quantity
            HAVING current_stock > 0 OR sold_qty > 0 ORDER BY sold_qty DESC
        """), {"start": start}).fetchall()
        result=[]
        for r in rows:
            sold=int(r.sold_qty or 0); stock=int(r.current_stock or 0)
            vel=round(sold/days_param,3); dl=round(stock/vel,1) if vel>0 else None
            result.append({"product_id":r.product_id,"name":r.name,"category":r.category or "",
                "current_stock":stock,"sold_in_period":sold,"velocity_per_day":vel,
                "days_of_stock_left":dl,
                "urgency":"Critical" if dl is not None and dl<=7 else
                           "Low"     if dl is not None and dl<=14 else
                           "OK"      if dl is not None else "No Sales"})
        order={"Critical":0,"Low":1,"OK":2,"No Sales":3}
        result.sort(key=lambda x:(order[x["urgency"]],x["days_of_stock_left"] or 9999))
        return jsonify(success=True, days=days_param, rows=result)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/reorder-needed")
@login_required
def api_reorder_needed():
    err = _admin_only()
    if err: return err
    try:
        rows = db.session.execute(text("""
            SELECT p.product_id, p.name, p.category, p.quantity, p.reorder_level,
                   (p.reorder_level - p.quantity) AS shortage,
                   s.company AS supplier_name, s.phone AS supplier_phone, s.email AS supplier_email
            FROM products p LEFT JOIN suppliers s ON s.supplier_id=p.supplier_id
            WHERE p.quantity <= p.reorder_level ORDER BY shortage DESC
        """)).fetchall()
        data=[{"product_id":r.product_id,"name":r.name,"category":r.category or "",
               "current_stock":int(r.quantity or 0),"reorder_level":int(r.reorder_level or 0),
               "shortage":int(r.shortage or 0),"supplier_name":r.supplier_name or "",
               "supplier_phone":r.supplier_phone or "","supplier_email":r.supplier_email or ""}
              for r in rows]
        return jsonify(success=True, rows=data, total_items=len(data))
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/return-analysis")
@login_required
def api_return_analysis():
    err = _admin_only()
    if err: return err
    s, e = _date_range()
    try:
        by_product = db.session.execute(text("""
            SELECT r.product_id, si.product_name,
                   SUM(r.quantity) AS returned_qty, SUM(r.refund_amount) AS total_refund, r.reason
            FROM returns r JOIN sales_items si ON si.product_id=r.product_id
            WHERE DATE(r.date) BETWEEN :s AND :e
            GROUP BY r.product_id, si.product_name, r.reason
            ORDER BY total_refund DESC LIMIT 20
        """), {"s":s,"e":e}).fetchall()

        by_payment = db.session.execute(text("""
            SELECT sm.payment_mode, COUNT(r.return_id) AS return_count,
                   SUM(r.refund_amount) AS total_refund
            FROM returns r JOIN sales_master sm ON r.sale_id=sm.sale_id
            WHERE DATE(r.date) BETWEEN :s AND :e
            GROUP BY sm.payment_mode ORDER BY total_refund DESC
        """), {"s":s,"e":e}).fetchall()

        totals = db.session.execute(text("""
            SELECT COUNT(*) AS total_returns, SUM(r.refund_amount) AS total_refund,
                   SUM(r.quantity) AS total_qty
            FROM returns r WHERE DATE(r.date) BETWEEN :s AND :e
        """), {"s":s,"e":e}).fetchone()

        total_sales_val = db.session.execute(text(
            "SELECT IFNULL(SUM(grand_total),0) FROM sales_master WHERE DATE(date) BETWEEN :s AND :e"
        ), {"s":s,"e":e}).scalar() or 0.0

        refund_total = float(totals.total_refund or 0)
        return jsonify(success=True,
            summary={"total_returns":int(totals.total_returns or 0),"total_refund":round(refund_total,2),
                     "total_qty":int(totals.total_qty or 0),
                     "refund_to_sales_pct":round(refund_total/float(total_sales_val)*100,2)
                         if total_sales_val else 0.0},
            by_product=[{"product_id":r.product_id,"product":r.product_name,
                "returned_qty":int(r.returned_qty or 0),"total_refund":round(float(r.total_refund or 0),2),
                "reason":r.reason or ""} for r in by_product],
            by_payment=[{"payment_mode":r.payment_mode or "Unknown","return_count":int(r.return_count or 0),
                "total_refund":round(float(r.total_refund or 0),2)} for r in by_payment])
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


# ══════════════════════════════════════════════════════════════
# EXPORT  (Excel, GSTR-1, Reorder, PDF)
# ══════════════════════════════════════════════════════════════

@reports_bp.route("/api/export-excel")
@login_required
def api_export_excel():
    err = _admin_only()
    if err: return err
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment

    report = request.args.get("report", "sales")
    s, e   = _date_range()

    def style_ws(ws, color="1D9E75"):
        fill = PatternFill("solid", fgColor=color)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ml+4, 40)

    try:
        if report == "sales":
            rows = db.session.execute(text(f"""
                SELECT sm.sale_id, DATE(sm.date) AS date, sm.customer_name,
                       sm.sold_by, sm.payment_mode, si.product_name, si.category,
                       si.quantity, si.mrp, si.effective_total,
                       IFNULL(p.cost_price,0) AS cost_price, ({_DISC_SQL}) AS discount_amt
                FROM sales_master sm JOIN sales_items si ON si.sale_id=sm.sale_id
                LEFT JOIN products p ON si.product_id=p.product_id
                WHERE DATE(sm.date) BETWEEN :s AND :e ORDER BY sm.date DESC
            """), {"s":s,"e":e}).fetchall()
            df=pd.DataFrame([dict(r._mapping) for r in rows]); fname=f"sales_{s}_to_{e}.xlsx"

        elif report == "profit":
            rows = db.session.execute(text(f"""
                SELECT si.product_name, si.category, SUM(si.quantity) AS qty_sold,
                       IFNULL(SUM(si.effective_total),0) AS revenue,
                       IFNULL(SUM({_DISC_SQL}),0) AS discount,
                       IFNULL(SUM(IFNULL(p.cost_price,0)*si.quantity),0) AS cogs,
                       IFNULL(SUM({_PROFIT_SQL}),0) AS net_profit
                FROM sales_items si LEFT JOIN products p ON si.product_id=p.product_id
                JOIN sales_master sm ON si.sale_id=sm.sale_id
                WHERE DATE(sm.date) BETWEEN :s AND :e
                GROUP BY si.product_id, si.product_name, si.category ORDER BY revenue DESC
            """), {"s":s,"e":e}).fetchall()
            df=pd.DataFrame([dict(r._mapping) for r in rows]); fname=f"profit_{s}_to_{e}.xlsx"

        elif report == "gst":
            rows = db.session.execute(text(f"""
                SELECT si.product_name, si.category, IFNULL(p.gst,0) AS gst_rate,
                       SUM(si.quantity) AS qty,
                       IFNULL(SUM(si.mrp*si.quantity - ({_DISC_SQL})),0) AS taxable_amt
                FROM sales_items si LEFT JOIN products p ON si.product_id=p.product_id
                JOIN sales_master sm ON si.sale_id=sm.sale_id
                WHERE DATE(sm.date) BETWEEN :s AND :e
                GROUP BY si.product_id, si.product_name, si.category, p.gst ORDER BY taxable_amt DESC
            """), {"s":s,"e":e}).fetchall()
            df=pd.DataFrame([dict(r._mapping) for r in rows])
            if not df.empty:
                df["cgst"]=df.apply(lambda r:round(r["taxable_amt"]*(r["gst_rate"]/2)/100,2),axis=1)
                df["sgst"]=df["cgst"]; df["total_gst"]=df["cgst"]+df["sgst"]
            fname=f"gst_{s}_to_{e}.xlsx"

        elif report == "dead-stock":
            cutoff=(date.today()-timedelta(days=30)).isoformat()
            rows=db.session.execute(text("""
                SELECT p.product_id, p.name, p.category, p.quantity, p.reorder_level,
                       IFNULL(p.cost_price,0) AS cost_price,
                       p.quantity*IFNULL(p.cost_price,0) AS stock_value
                FROM products p WHERE p.quantity > 0
                  AND p.product_id NOT IN (
                      SELECT DISTINCT si.product_id FROM sales_items si
                      JOIN sales_master sm ON si.sale_id=sm.sale_id WHERE DATE(sm.date) >= :cutoff)
                ORDER BY stock_value DESC
            """),{"cutoff":cutoff}).fetchall()
            df=pd.DataFrame([dict(r._mapping) for r in rows]); fname="dead_stock.xlsx"

        elif report == "staff":
            rows=db.session.execute(text(f"""
                SELECT sm.sold_by AS employee, COUNT(DISTINCT sm.sale_id) AS bill_count,
                       IFNULL(SUM(sm.grand_total),0) AS total_revenue,
                       IFNULL(SUM({_PROFIT_SQL}),0)  AS total_profit,
                       IFNULL(SUM({_DISC_SQL}),0)     AS total_discount
                FROM sales_master sm JOIN sales_items si ON si.sale_id=sm.sale_id
                LEFT JOIN products p ON si.product_id=p.product_id
                WHERE DATE(sm.date) BETWEEN :s AND :e
                GROUP BY sm.sold_by ORDER BY total_revenue DESC
            """),{"s":s,"e":e}).fetchall()
            df=pd.DataFrame([dict(r._mapping) for r in rows])
            if not df.empty and "bill_count" in df.columns:
                df["avg_bill"]=(df["total_revenue"]/df["bill_count"]).round(2)
            fname=f"staff_performance_{s}_to_{e}.xlsx"

        elif report == "customer-sales":
            rows=db.session.execute(text("""
                SELECT sm.customer_name, sm.customer_phone AS phone,
                       COUNT(DISTINCT sm.sale_id) AS order_count, SUM(sm.grand_total) AS total_spend,
                       MIN(DATE(sm.date)) AS first_purchase, MAX(DATE(sm.date)) AS last_purchase
                FROM sales_master sm WHERE DATE(sm.date) BETWEEN :s AND :e
                GROUP BY sm.customer_name, sm.customer_phone ORDER BY total_spend DESC
            """),{"s":s,"e":e}).fetchall()
            df=pd.DataFrame([dict(r._mapping) for r in rows])
            if not df.empty: df["avg_order"]=(df["total_spend"]/df["order_count"]).round(2)
            fname=f"customer_sales_{s}_to_{e}.xlsx"

        else:
            return jsonify(success=False, message=f"Unknown report '{report}'"), 400

        if df.empty:
            return jsonify(success=False, message="No data found."), 404

        buf=io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Report")
            style_ws(writer.sheets["Report"])
        buf.seek(0)

        xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if _wants_email():
            return _email_report(
                filename=fname,
                mimetype=xlsx_mime,
                file_bytes=buf.getvalue(),
                subject=f"IMS Report: {report.title()} ({s} to {e})",
                body=(f"Attached: {report.title()} report for the period "
                      f"{s} to {e}.\n\nGenerated by IMS."),
            )

        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype=xlsx_mime)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500
# ══════════════════════════════════════════════════════════════
# REORDER AGENT — pure Python, zero external dependencies
# Reads DB history → computes smart reorder quantities
# ══════════════════════════════════════════════════════════════

from math import ceil, sqrt

@reports_bp.route("/api/reorder-agent")
@login_required
def api_reorder_agent():
    err = _admin_only()
    if err: return err

    try:
        lookback_days = int(request.args.get("days", 60))
        safety_factor = float(request.args.get("safety_factor", 1.5))
        start = (date.today() - timedelta(days=lookback_days)).isoformat()

        # ── 1. Sales velocity per product (from your own history) ──
        velocity_rows = db.session.execute(text("""
            SELECT
                p.product_id,
                p.name,
                p.category,
                p.quantity            AS current_stock,
                IFNULL(p.reorder_level, 0)  AS reorder_level,
                IFNULL(p.cost_price, 0)     AS cost_price,
                IFNULL(p.mrp, 0)            AS mrp,
                s.supplier_id,
                s.company             AS supplier_name,
                s.phone               AS supplier_phone,
                s.email               AS supplier_email,
                IFNULL(SUM(si.quantity), 0) AS sold_in_period,
                COUNT(DISTINCT DATE(sm.date)) AS active_days
            FROM products p
            LEFT JOIN suppliers s ON s.supplier_id = p.supplier_id
            LEFT JOIN sales_items si ON si.product_id = p.product_id
            LEFT JOIN sales_master sm
                ON  si.sale_id = sm.sale_id
                AND DATE(sm.date) >= :start
            GROUP BY
                p.product_id, p.name, p.category,
                p.quantity, p.reorder_level, p.cost_price, p.mrp,
                s.supplier_id, s.company, s.phone, s.email
        """), {"start": start}).fetchall()

        # ── 2. Day-level sales for std-dev (demand variability) ────
        daily_rows = db.session.execute(text("""
            SELECT
                si.product_id,
                DATE(sm.date)       AS day,
                SUM(si.quantity)    AS qty
            FROM sales_items si
            JOIN sales_master sm ON si.sale_id = sm.sale_id
            WHERE DATE(sm.date) >= :start
            GROUP BY si.product_id, DATE(sm.date)
        """), {"start": start}).fetchall()

        # ── 3. Past purchase orders (lead time learning) ────────────
        # If you have a purchases table use it; else we default to 7 days
        try:
            lead_rows = db.session.execute(text("""
                SELECT product_id,
                       AVG(DATEDIFF(received_date, order_date)) AS avg_lead_days
                FROM purchase_orders
                WHERE order_date >= :start
                GROUP BY product_id
            """), {"start": start}).fetchall()
            lead_map = {r.product_id: float(r.avg_lead_days or 7) for r in lead_rows}
        except Exception:
            lead_map = {}   # table may not exist yet

        # ── 4. Build daily demand map for std-dev ──────────────────
        from collections import defaultdict
        daily_map = defaultdict(list)
        for r in daily_rows:
            daily_map[r.product_id].append(float(r.qty or 0))

        # ── 5. Core reorder logic per product ──────────────────────
        actions   = []
        summaries = {"urgent": 0, "soon": 0, "watch": 0, "ok": 0}

        for r in velocity_rows:
            pid          = r.product_id
            stock        = int(r.current_stock or 0)
            reorder_lvl  = int(r.reorder_level or 0)
            sold         = float(r.sold_in_period or 0)
            cost         = float(r.cost_price or 0)
            active_days  = max(int(r.active_days or 1), 1)

            # Average daily demand (only on days it actually sold)
            avg_daily = sold / lookback_days  # conservative: spread over full period

            # Demand std-dev for safety stock calculation
            day_qtys = daily_map.get(pid, [0])
            if len(day_qtys) > 1:
                mean = sum(day_qtys) / len(day_qtys)
                variance = sum((x - mean) ** 2 for x in day_qtys) / len(day_qtys)
                std_dev = sqrt(variance)
            else:
                std_dev = avg_daily * 0.3   # assume 30% variability if little data

            # Lead time (days supplier takes to deliver)
            lead_time = lead_map.get(pid, 7)

            # Safety stock = Z * std_dev * sqrt(lead_time)
            # Z=1.65 → 95% service level
            Z = 1.65
            safety_stock = ceil(Z * std_dev * sqrt(lead_time))

            # Reorder point: demand during lead time + safety stock
            reorder_point = ceil(avg_daily * lead_time + safety_stock)

            # Economic Order Quantity (Wilson formula)
            # EOQ = sqrt(2 * D * S / H)
            # D = annual demand, S = ordering cost (assume ₹50), H = holding cost (20% of cost)
            annual_demand = avg_daily * 365
            ordering_cost = 50.0
            holding_cost  = max(cost * 0.20, 1.0)
            if annual_demand > 0 and holding_cost > 0:
                eoq = ceil(sqrt(2 * annual_demand * ordering_cost / holding_cost))
            else:
                eoq = max(reorder_lvl, 10)

            # Days of stock remaining
            days_left = round(stock / avg_daily, 1) if avg_daily > 0 else None

            # ── Decision ────────────────────────────────────────────
            if stock == 0:
                urgency = "urgent"
                reason  = "Out of stock — immediate order required"
                order_qty = max(eoq, reorder_lvl * 2)
            elif stock <= safety_stock:
                urgency = "urgent"
                reason  = f"Below safety stock ({safety_stock} units). {days_left or '?'} days remaining"
                order_qty = eoq
            elif stock <= reorder_point:
                urgency = "soon"
                reason  = f"At reorder point. ~{days_left or '?'} days left before stockout"
                order_qty = eoq
            elif reorder_lvl > 0 and stock <= reorder_lvl:
                urgency = "watch"
                reason  = f"Below reorder level ({reorder_lvl}). Monitor closely"
                order_qty = max(ceil(eoq * 0.5), reorder_lvl - stock)
            else:
                urgency = "ok"
                reason  = f"Adequate stock. ~{days_left or '?'} days remaining"
                order_qty = 0

            summaries[urgency] += 1

            # Only include products that need action or are being watched
            if urgency == "ok":
                continue

            # Trend: compare first-half vs second-half of the period
            mid = start
            first_half_sold = db.session.execute(text("""
                SELECT IFNULL(SUM(si.quantity), 0)
                FROM sales_items si JOIN sales_master sm ON si.sale_id=sm.sale_id
                WHERE si.product_id=:pid
                  AND DATE(sm.date) BETWEEN :s AND :m
            """), {
                "pid": pid,
                "s": start,
                "m": (date.fromisoformat(start) + timedelta(days=lookback_days//2)).isoformat()
            }).scalar() or 0

            second_half_sold = sold - float(first_half_sold)
            if first_half_sold > 0:
                trend_pct = round((second_half_sold - first_half_sold) / first_half_sold * 100, 1)
            else:
                trend_pct = 0.0

            trend_label = (
                "rising" if trend_pct > 15
                else "falling" if trend_pct < -15
                else "stable"
            )

            # Adjust order qty for trend
            if trend_label == "rising":
                order_qty = ceil(order_qty * 1.2)
            elif trend_label == "falling":
                order_qty = ceil(order_qty * 0.8)

            actions.append({
                "product_id":      pid,
                "name":            r.name,
                "category":        r.category or "",
                "current_stock":   stock,
                "reorder_point":   reorder_point,
                "safety_stock":    safety_stock,
                "eoq":             eoq,
                "recommended_qty": int(order_qty),
                "estimated_cost":  round(order_qty * cost, 2),
                "avg_daily_demand":round(avg_daily, 2),
                "std_dev_demand":  round(std_dev, 2),
                "lead_time_days":  lead_time,
                "days_left":       days_left,
                "trend":           trend_label,
                "trend_pct":       trend_pct,
                "urgency":         urgency,
                "reason":          reason,
                "supplier_id":     r.supplier_id,
                "supplier_name":   r.supplier_name or "Unknown",
                "supplier_phone":  r.supplier_phone or "",
                "supplier_email":  r.supplier_email or "",
            })

        # ── 6. Sort: urgent first, then by days_left ascending ─────
        urgency_order = {"urgent": 0, "soon": 1, "watch": 2}
        actions.sort(key=lambda x: (
            urgency_order[x["urgency"]],
            x["days_left"] if x["days_left"] is not None else 999
        ))

        # ── 7. Group by supplier for purchase order view ────────────
        supplier_orders = defaultdict(lambda: {
            "supplier_name": "", "supplier_phone": "", "supplier_email": "",
            "items": [], "total_cost": 0.0
        })
        for a in actions:
            sid = a["supplier_id"] or "unknown"
            supplier_orders[sid]["supplier_name"]  = a["supplier_name"]
            supplier_orders[sid]["supplier_phone"] = a["supplier_phone"]
            supplier_orders[sid]["supplier_email"] = a["supplier_email"]
            supplier_orders[sid]["items"].append({
                "product_id":  a["product_id"],
                "name":        a["name"],
                "qty":         a["recommended_qty"],
                "unit_cost":   a["estimated_cost"] / a["recommended_qty"] if a["recommended_qty"] else 0,
                "total_cost":  a["estimated_cost"],
                "urgency":     a["urgency"],
            })
            supplier_orders[sid]["total_cost"] += a["estimated_cost"]

        supplier_list = [
            {**v, "supplier_id": k, "total_cost": round(v["total_cost"], 2)}
            for k, v in supplier_orders.items()
        ]
        supplier_list.sort(key=lambda x: -x["total_cost"])

        return jsonify(
            success=True,
            generated_at=date.today().isoformat(),
            lookback_days=lookback_days,
            summary=summaries,
            total_items_flagged=len(actions),
            total_estimated_spend=round(sum(a["estimated_cost"] for a in actions), 2),
            actions=actions,
            by_supplier=supplier_list,
        )

    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500

@reports_bp.route("/api/export-gstr1")
@login_required
def api_export_gstr1():
    err = _admin_only()
    if err: return err
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment

    month=request.args.get("month", date.today().strftime("%m"))
    year =request.args.get("year",  date.today().strftime("%Y"))

    def style_ws(ws, color="1D9E75"):
        fill=PatternFill("solid",fgColor=color)
        for cell in ws[1]:
            cell.font=Font(bold=True,color="FFFFFF"); cell.fill=fill
            cell.alignment=Alignment(horizontal="center")
        for col in ws.columns:
            ml=max((len(str(c.value or "")) for c in col),default=10)
            ws.column_dimensions[col[0].column_letter].width=min(ml+4,35)

    try:
        slab_rows=db.session.execute(text(f"""
            SELECT IFNULL(p.gst,0) AS gst_rate, COUNT(DISTINCT sm.sale_id) AS txns,
                   SUM(si.quantity) AS total_qty,
                   IFNULL(SUM(si.mrp*si.quantity-({_DISC_SQL})),0) AS taxable_amt
            FROM sales_items si LEFT JOIN products p ON si.product_id=p.product_id
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE MONTH(sm.date)=:month AND YEAR(sm.date)=:year
            GROUP BY p.gst ORDER BY p.gst
        """),{"month":int(month),"year":int(year)}).fetchall()

        detail_rows=db.session.execute(text(f"""
            SELECT si.product_name, si.category, IFNULL(p.gst,0) AS gst_rate,
                   SUM(si.quantity) AS qty,
                   IFNULL(SUM(si.mrp*si.quantity-({_DISC_SQL})),0) AS taxable_amt
            FROM sales_items si LEFT JOIN products p ON si.product_id=p.product_id
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE MONTH(sm.date)=:month AND YEAR(sm.date)=:year
            GROUP BY si.product_id, si.product_name, si.category, p.gst ORDER BY taxable_amt DESC
        """),{"month":int(month),"year":int(year)}).fetchall()

        def to_slab(rows):
            data=[]
            for r in rows:
                rate=float(r.gst_rate or 0); tax=max(float(r.taxable_amt or 0),0.0)
                cgst=round(tax*(rate/2)/100,2); sgst=cgst
                data.append({"GST Rate (%)":rate,"Transactions":int(r.txns or 0),
                              "Total Qty":int(r.total_qty or 0),"Taxable Amt":round(tax,2),
                              "CGST":cgst,"SGST":sgst,"Total Tax":round(cgst+sgst,2)})
            return pd.DataFrame(data)

        def to_detail(rows):
            data=[]
            for r in rows:
                rate=float(r.gst_rate or 0); tax=max(float(r.taxable_amt or 0),0.0)
                cgst=round(tax*(rate/2)/100,2); sgst=cgst
                data.append({"Product":r.product_name,"Category":r.category or "Other",
                              "GST Rate (%)":rate,"Qty Sold":int(r.qty or 0),"Taxable Amt":round(tax,2),
                              "CGST":cgst,"SGST":sgst,"Total Tax":round(cgst+sgst,2)})
            return pd.DataFrame(data)

        buf=io.BytesIO()
        with pd.ExcelWriter(buf,engine="openpyxl") as writer:
            to_slab(slab_rows).to_excel(writer,index=False,sheet_name="Summary")
            to_detail(detail_rows).to_excel(writer,index=False,sheet_name="Product Detail")
            for sh in ["Summary","Product Detail"]: style_ws(writer.sheets[sh])
        buf.seek(0)

        xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        fname = f"GSTR1_{year}_{month}.xlsx"
        if _wants_email():
            return _email_report(
                filename=fname,
                mimetype=xlsx_mime,
                file_bytes=buf.getvalue(),
                subject=f"IMS GSTR-1 Report ({month}/{year})",
                body=f"Attached: GSTR-1 filing data for {month}/{year}.\n\nGenerated by IMS.",
            )

        return send_file(buf,as_attachment=True,download_name=fname, mimetype=xlsx_mime)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/export-reorder")
@login_required
def api_export_reorder():
    err = _admin_only()
    if err: return err
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    try:
        rows=db.session.execute(text("""
            SELECT p.product_id AS "Product ID", p.name AS "Product Name",
                   p.category AS "Category", p.quantity AS "Current Stock",
                   p.reorder_level AS "Reorder Level",
                   (p.reorder_level-p.quantity) AS "Shortage",
                   s.company AS "Supplier", s.phone AS "Supplier Phone",
                   s.email AS "Supplier Email"
            FROM products p LEFT JOIN suppliers s ON s.supplier_id=p.supplier_id
            WHERE p.quantity <= p.reorder_level
            ORDER BY (p.reorder_level-p.quantity) DESC
        """)).fetchall()
        df=pd.DataFrame([dict(r._mapping) for r in rows])
        buf=io.BytesIO()
        with pd.ExcelWriter(buf,engine="openpyxl") as writer:
            df.to_excel(writer,index=False,sheet_name="Reorder List")
            ws=writer.sheets["Reorder List"]
            fill=PatternFill("solid",fgColor="E85D24")
            for cell in ws[1]:
                cell.font=Font(bold=True,color="FFFFFF"); cell.fill=fill
                cell.alignment=Alignment(horizontal="center")
            for col in ws.columns:
                ml=max((len(str(c.value or "")) for c in col),default=10)
                ws.column_dimensions[col[0].column_letter].width=min(ml+4,35)
        buf.seek(0)

        xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if _wants_email():
            return _email_report(
                filename="reorder_list.xlsx",
                mimetype=xlsx_mime,
                file_bytes=buf.getvalue(),
                subject=f"IMS Reorder List — {date.today().isoformat()}",
                body=f"Attached: current reorder list as of {date.today().isoformat()}.\n\nGenerated by IMS.",
            )

        return send_file(buf,as_attachment=True,download_name="reorder_list.xlsx",
                         mimetype=xlsx_mime)
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


# ══════════════════════════════════════════════════════════════
# YEAR AUDIT  (Financial Year: April → March)
# ══════════════════════════════════════════════════════════════

@reports_bp.route("/api/year-audit")
@login_required
def api_year_audit():
    """
    Returns per-product accumulated figures for a given financial year
    (April 1 → March 31).

    Query params:
      fy   – financial year start, e.g. 2024  → 2024-04-01 … 2025-03-31
             defaults to the current FY
    """
    err = _admin_only()
    if err: return err

    today_d = date.today()
    # default FY start year: if current month < April, FY started previous year
    default_fy = today_d.year if today_d.month >= 4 else today_d.year - 1
    fy_start_year = request.args.get("fy", default_fy, type=int)

    fy_from = f"{fy_start_year}-04-01"
    fy_to   = f"{fy_start_year + 1}-03-31"

    try:
        rows = db.session.execute(text(f"""
            SELECT
                si.product_id,
                si.product_name,
                IFNULL(si.category, 'Other')                       AS category,
                IFNULL(p.gst, 0)                                   AS gst_rate,
                IFNULL(p.cost_price, 0)                            AS cost_price,

                /* total units sold in FY */
                SUM(si.quantity)                                    AS total_qty,

                /* accumulated cost  = qty * cost_price  (current cost_price) */
                SUM(si.quantity * IFNULL(p.cost_price, 0))         AS acc_cost,

                /* taxable value after discount */
                SUM(
                    si.mrp * si.quantity
                    - CASE si.discount_type
                        WHEN 'Flat' THEN IFNULL(si.discount_value, 0)
                        ELSE si.mrp * si.quantity * IFNULL(si.discount_value, 0) / 100.0
                      END
                )                                                   AS taxable_value,

                /* accumulated GST = taxable * gst_rate% */
                SUM(
                    IFNULL(p.gst, 0) / 100.0 * (
                        si.mrp * si.quantity
                        - CASE si.discount_type
                            WHEN 'Flat' THEN IFNULL(si.discount_value, 0)
                            ELSE si.mrp * si.quantity * IFNULL(si.discount_value, 0) / 100.0
                          END
                    )
                )                                                   AS acc_gst,

                /* accumulated discount */
                SUM(
                    CASE si.discount_type
                        WHEN 'Flat' THEN IFNULL(si.discount_value, 0)
                        ELSE si.mrp * si.quantity * IFNULL(si.discount_value, 0) / 100.0
                    END
                )                                                   AS acc_discount,

                /* accumulated profit =
                      (sell_price - discount - gst) - cost
                   = taxable_value + gst - gst - cost
                   = effective_total - cost  (simpler, same result) */
                SUM(si.effective_total)
                - SUM(si.quantity * IFNULL(p.cost_price, 0))       AS acc_profit

            FROM sales_items si
            LEFT JOIN products p    ON si.product_id = p.product_id
            JOIN  sales_master sm   ON si.sale_id    = sm.sale_id
            WHERE DATE(sm.date) BETWEEN :fy_from AND :fy_to
            GROUP BY si.product_id, si.product_name, si.category, p.gst, p.cost_price
            ORDER BY acc_profit DESC
        """), {"fy_from": fy_from, "fy_to": fy_to}).fetchall()

        result = []
        grand_cost = grand_profit = grand_gst = grand_discount = grand_qty = 0.0

        for r in rows:
            qty        = int(r.total_qty    or 0)
            acc_cost   = round(float(r.acc_cost    or 0), 2)
            acc_profit = round(float(r.acc_profit  or 0), 2)
            acc_gst    = round(float(r.acc_gst     or 0), 2)
            acc_disc   = round(float(r.acc_discount or 0), 2)
            taxable    = round(float(r.taxable_value or 0), 2)
            gst_rate   = float(r.gst_rate  or 0)
            cgst       = round(acc_gst / 2, 2)
            sgst       = round(acc_gst / 2, 2)
            margin     = round(acc_profit / taxable * 100, 2) if taxable else 0.0

            grand_cost     += acc_cost
            grand_profit   += acc_profit
            grand_gst      += acc_gst
            grand_discount += acc_disc
            grand_qty      += qty

            result.append({
                "product_id":   r.product_id,
                "product_name": r.product_name,
                "category":     r.category,
                "gst_rate":     gst_rate,
                "cost_price":   float(r.cost_price or 0),
                "qty_sold":     qty,
                "acc_cost":     acc_cost,
                "acc_discount": acc_disc,
                "taxable":      taxable,
                "acc_gst":      acc_gst,
                "cgst":         cgst,
                "sgst":         sgst,
                "acc_profit":   acc_profit,
                "margin_pct":   margin,
            })

        return jsonify(
            success   = True,
            fy_label  = f"FY {fy_start_year}–{fy_start_year + 1}",
            fy_from   = fy_from,
            fy_to     = fy_to,
            rows      = result,
            summary   = {
                "total_qty":      int(grand_qty),
                "total_cost":     round(grand_cost, 2),
                "total_discount": round(grand_discount, 2),
                "total_gst":      round(grand_gst, 2),
                "total_profit":   round(grand_profit, 2),
                "cgst":           round(grand_gst / 2, 2),
                "sgst":           round(grand_gst / 2, 2),
            }
        )

    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/export-year-audit")
@login_required
def api_export_year_audit():
    """Excel export for Year Audit tab."""
    err = _admin_only()
    if err: return err

    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today_d = date.today()
    default_fy = today_d.year if today_d.month >= 4 else today_d.year - 1
    fy_start_year = request.args.get("fy", default_fy, type=int)
    fy_from = f"{fy_start_year}-04-01"
    fy_to   = f"{fy_start_year + 1}-03-31"
    fy_label = f"FY{fy_start_year}-{fy_start_year + 1}"

    try:
        rows = db.session.execute(text(f"""
            SELECT
                si.product_id,
                si.product_name,
                IFNULL(si.category,'Other') AS category,
                IFNULL(p.gst,0)             AS gst_rate,
                SUM(si.quantity)            AS total_qty,
                SUM(si.quantity * IFNULL(p.cost_price,0)) AS acc_cost,
                SUM(
                    IFNULL(p.gst,0)/100.0 * (
                        si.mrp*si.quantity
                        - CASE si.discount_type
                            WHEN 'Flat' THEN IFNULL(si.discount_value,0)
                            ELSE si.mrp*si.quantity*IFNULL(si.discount_value,0)/100.0
                          END
                    )
                ) AS acc_gst,
                SUM(
                    CASE si.discount_type
                        WHEN 'Flat' THEN IFNULL(si.discount_value,0)
                        ELSE si.mrp*si.quantity*IFNULL(si.discount_value,0)/100.0
                    END
                ) AS acc_discount,
                SUM(si.effective_total)
                - SUM(si.quantity*IFNULL(p.cost_price,0)) AS acc_profit
            FROM sales_items si
            LEFT JOIN products p  ON si.product_id=p.product_id
            JOIN  sales_master sm ON si.sale_id=sm.sale_id
            WHERE DATE(sm.date) BETWEEN :fy_from AND :fy_to
            GROUP BY si.product_id,si.product_name,si.category,p.gst
            ORDER BY acc_profit DESC
        """), {"fy_from": fy_from, "fy_to": fy_to}).fetchall()

        data = []
        for r in rows:
            gst = round(float(r.acc_gst or 0), 2)
            data.append({
                "Product ID":    r.product_id,
                "Product Name":  r.product_name,
                "Category":      r.category,
                "GST Rate (%)":  float(r.gst_rate or 0),
                "Qty Sold":      int(r.total_qty or 0),
                "Acc. Cost (₹)": round(float(r.acc_cost or 0), 2),
                "Acc. Discount (₹)": round(float(r.acc_discount or 0), 2),
                "CGST (₹)":     round(gst / 2, 2),
                "SGST (₹)":     round(gst / 2, 2),
                "Total GST (₹)": gst,
                "Acc. Profit (₹)": round(float(r.acc_profit or 0), 2),
            })

        df  = pd.DataFrame(data)
        buf = io.BytesIO()

        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Year Audit", startrow=2)
            ws = writer.sheets["Year Audit"]

            # Title row
            ws.merge_cells("A1:K1")
            title_cell = ws["A1"]
            title_cell.value = f"Lalbagh Enterprise – Year Audit {fy_label}  ({fy_from} to {fy_to})"
            title_cell.font  = Font(bold=True, size=13, color="FFFFFF")
            title_cell.fill  = PatternFill("solid", fgColor="1D9E75")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

            # Header row styling
            hdr_fill = PatternFill("solid", fgColor="2D3748")
            for cell in ws[3]:
                cell.font      = Font(bold=True, color="FFFFFF", size=9)
                cell.fill      = hdr_fill
                cell.alignment = Alignment(horizontal="center")

            # Column widths
            col_widths = [14, 38, 18, 12, 10, 16, 18, 14, 14, 14, 18]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # Totals row
            last = len(data) + 3
            ws.cell(row=last+1, column=1, value="TOTAL").font = Font(bold=True)
            num_cols = {5: "Qty Sold", 6: "Acc. Cost (₹)", 7: "Acc. Discount (₹)",
                        8: "CGST (₹)", 9: "SGST (₹)", 10: "Total GST (₹)", 11: "Acc. Profit (₹)"}
            for col_idx in num_cols:
                col_letter = get_column_letter(col_idx)
                total_cell = ws.cell(row=last+1, column=col_idx)
                total_cell.value = f"=SUM({col_letter}4:{col_letter}{last})"
                total_cell.font  = Font(bold=True)
                total_cell.fill  = PatternFill("solid", fgColor="E8F5E9")

        buf.seek(0)

        xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        fname = f"year_audit_{fy_label}.xlsx"
        if _wants_email():
            return _email_report(
                filename=fname,
                mimetype=xlsx_mime,
                file_bytes=buf.getvalue(),
                subject=f"IMS Year Audit — {fy_label} ({fy_from} to {fy_to})",
                body=(f"Attached: Year Audit report for {fy_label} "
                      f"({fy_from} to {fy_to}).\n\nGenerated by IMS."),
            )

        return send_file(
            buf,
            as_attachment=True,
            download_name=fname,
            mimetype=xlsx_mime
        )

    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, message=str(ex)), 500


@reports_bp.route("/api/export-pdf")
@login_required
def api_export_pdf():
    err = _admin_only()
    if err: return err

    import matplotlib; matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, Image as RLImage,
                                     PageBreak, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors as rl_colors

    s, e     = _date_range()
    tmp_imgs = []
    out_path = os.path.join(current_app.config["INVOICE_DIR"], "erp_report.pdf")
    PAGE_W   = A4[0] - 36
    TEAL     = rl_colors.HexColor("#1D9E75")
    GREY     = rl_colors.HexColor("#F5F5F5")
    AMBER    = rl_colors.HexColor("#F39C12")

    def hdr_style(bg=TEAL):
        return TableStyle([("BACKGROUND",(0,0),(-1,0),bg),("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
            ("GRID",(0,0),(-1,-1),0.25,rl_colors.lightgrey),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.white,GREY]),
            ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3)])

    def save_fig(fig):
        t=tempfile.NamedTemporaryFile(suffix=".png",delete=False)
        fig.savefig(t.name,bbox_inches="tight",dpi=140,facecolor="white")
        tmp_imgs.append(t.name); plt.close(fig); return t.name

    try:
        styles=getSampleStyleSheet()
        title_st=ParagraphStyle("T",parent=styles["Title"],fontSize=16,textColor=TEAL,spaceAfter=4)
        h2_st   =ParagraphStyle("H2",parent=styles["Heading2"],fontSize=11,textColor=TEAL,spaceBefore=10,spaceAfter=4)
        normal  =styles["Normal"]; normal.fontSize=8

        doc=SimpleDocTemplate(out_path,pagesize=A4,rightMargin=18,leftMargin=18,topMargin=18,bottomMargin=18)
        story=[]
        company=current_app.config.get("COMPANY_NAME","LALBAGH ENTERPRISE")
        story+=[Paragraph(company,title_st),Paragraph("Business Analytics Report",styles["Title"]),
                Paragraph(f"Period: {s} to {e}",normal),
                Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | By: {current_user.username}",normal),
                HRFlowable(width="100%",thickness=1,color=TEAL,spaceAfter=8)]

        # KPI fetch
        ts=float(db.session.execute(text("SELECT IFNULL(SUM(grand_total),0) FROM sales_master WHERE DATE(date) BETWEEN :s AND :e"),{"s":s,"e":e}).scalar() or 0)
        tp=float(db.session.execute(text(f"SELECT IFNULL(SUM({_PROFIT_SQL}),0) FROM sales_items si LEFT JOIN products p ON si.product_id=p.product_id JOIN sales_master sm ON si.sale_id=sm.sale_id WHERE DATE(sm.date) BETWEEN :s AND :e"),{"s":s,"e":e}).scalar() or 0)
        bills=int(db.session.execute(text("SELECT COUNT(*) FROM sales_master WHERE DATE(date) BETWEEN :s AND :e"),{"s":s,"e":e}).scalar() or 0)
        ret=float(db.session.execute(text("SELECT IFNULL(SUM(refund_amount),0) FROM returns WHERE DATE(date) BETWEEN :s AND :e"),{"s":s,"e":e}).scalar() or 0)
        margin=round(tp/ts*100,2) if ts else 0.0; avg_bill=round(ts/bills,2) if bills else 0.0

        kd=[["Metric","Value"],["Total Sales",f"₹{ts:,.2f}"],["Net Profit",f"₹{tp:,.2f}"],
            ["Margin",f"{margin:.2f}%"],["Bills",str(bills)],["Avg Bill",f"₹{avg_bill:,.2f}"],["Returns",f"₹{ret:,.2f}"]]
        kt=Table(kd,colWidths=[PAGE_W*0.45,PAGE_W*0.55]); kt.setStyle(hdr_style())
        story+=[Paragraph("KPI Summary",h2_st),kt,Spacer(1,10)]

        # Monthly chart
        mr=db.session.execute(text(f"""
            SELECT DATE_FORMAT(sm.date,'%Y-%m') AS ym,
                   IFNULL(SUM(sm.grand_total),0) AS sales, IFNULL(SUM({_PROFIT_SQL}),0) AS profit
            FROM sales_master sm JOIN sales_items si ON si.sale_id=sm.sale_id
            LEFT JOIN products p ON si.product_id=p.product_id
            GROUP BY ym ORDER BY ym DESC LIMIT 12""")).fetchall()
        mr=list(reversed(mr))
        if mr:
            fig,ax=plt.subplots(figsize=(9,3.8))
            x=range(len(mr))
            ax.fill_between(x,[float(r[1]) for r in mr],alpha=0.1,color="#1D9E75")
            ax.plot(x,[float(r[1]) for r in mr],"o-",color="#1D9E75",lw=2,label="Sales")
            ax.plot(x,[float(r[2]) for r in mr],"s--",color="#27AE60",lw=1.5,label="Profit")
            ax.set_xticks(x); ax.set_xticklabels([r[0] for r in mr],rotation=40,ha="right",fontsize=7)
            ax.set_title("Monthly Sales & Profit",fontsize=10,pad=8)
            ax.legend(fontsize=8); ax.grid(axis="y",linestyle="--",alpha=0.3)
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v,_:f"₹{v:,.0f}"))
            story+=[Paragraph("Monthly Trend",h2_st),RLImage(save_fig(fig),width=PAGE_W,height=190),Spacer(1,8)]

        # Top products chart
        tp_rows=db.session.execute(text(f"""
            SELECT si.product_name, IFNULL(SUM(si.effective_total),0) AS revenue, IFNULL(SUM({_PROFIT_SQL}),0) AS profit
            FROM sales_items si LEFT JOIN products p ON si.product_id=p.product_id
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE DATE(sm.date) BETWEEN :s AND :e
            GROUP BY si.product_id, si.product_name ORDER BY revenue DESC LIMIT 10"""),{"s":s,"e":e}).fetchall()
        if tp_rows:
            names=[r[0][:20] for r in tp_rows]; rev_v=[float(r[1] or 0) for r in tp_rows]; prof_v=[float(r[2] or 0) for r in tp_rows]
            fig,ax=plt.subplots(figsize=(9,3.8))
            yi=range(len(names))
            ax.barh(yi,rev_v,height=0.4,label="Revenue",color="#1D9E75")
            ax.barh([y+0.4 for y in yi],prof_v,height=0.4,label="Profit",color="#27AE60")
            ax.set_yticks([y+0.2 for y in yi]); ax.set_yticklabels(names,fontsize=7)
            ax.set_title("Top 10 Products",fontsize=10,pad=8); ax.legend(fontsize=8)
            ax.grid(axis="x",linestyle="--",alpha=0.3)
            ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda v,_:f"₹{v:,.0f}"))
            story+=[Paragraph("Top Products",h2_st),RLImage(save_fig(fig),width=PAGE_W,height=190),Spacer(1,8)]

        story.append(PageBreak())

        # Profit table
        pr=db.session.execute(text(f"""
            SELECT si.product_name, SUM(si.quantity) AS qty,
                   IFNULL(SUM(si.effective_total),0) AS revenue,
                   IFNULL(SUM({_DISC_SQL}),0) AS disc,
                   IFNULL(SUM(IFNULL(p.cost_price,0)*si.quantity),0) AS cogs,
                   IFNULL(SUM({_PROFIT_SQL}),0) AS profit
            FROM sales_items si LEFT JOIN products p ON si.product_id=p.product_id
            JOIN sales_master sm ON si.sale_id=sm.sale_id
            WHERE DATE(sm.date) BETWEEN :s AND :e
            GROUP BY si.product_id, si.product_name ORDER BY profit DESC LIMIT 20"""),{"s":s,"e":e}).fetchall()
        story.append(Paragraph("Profit by Product (Top 20)",h2_st))
        prd=[["Product","Qty","Revenue","Discount","COGS","Net Profit","Margin%"]]
        for r in pr:
            rev=float(r.revenue or 0); prof=float(r.profit or 0); mg=round(prof/rev*100,1) if rev else 0.0
            prd.append([r.product_name[:28],int(r.qty or 0),f"₹{rev:,.0f}",f"₹{float(r.disc or 0):,.0f}",
                        f"₹{float(r.cogs or 0):,.0f}",f"₹{prof:,.0f}",f"{mg}%"])
        pt=Table(prd,colWidths=[PAGE_W*0.30,PAGE_W*0.07,PAGE_W*0.14,PAGE_W*0.13,PAGE_W*0.13,PAGE_W*0.14,PAGE_W*0.09])
        pt.setStyle(hdr_style()); story+=[pt,Spacer(1,10)]

        doc.build(story)

        pdf_fname = f"erp_report_{s}_to_{e}.pdf"
        if _wants_email():
            with open(out_path, "rb") as f:
                pdf_bytes = f.read()
            return _email_report(
                filename=pdf_fname,
                mimetype="application/pdf",
                file_bytes=pdf_bytes,
                subject=f"IMS Analytics Report ({s} to {e})",
                body=(f"Attached: full business analytics report (KPIs, trends, "
                      f"top products, profit by product) for {s} to {e}.\n\n"
                      f"Generated by IMS."),
            )

        return send_file(out_path,as_attachment=True,download_name=pdf_fname,
                         mimetype="application/pdf")
    finally:
        for fp in tmp_imgs:
            try: os.remove(fp)
            except: pass